import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_wb(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    #cell= sheet['a1']
    #cell=sheet.cell(1,1)

    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        new_price=cell.value * 0.9
        #print(new_price)
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=new_price

    # build a chart
    values=Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    # save file
    wb.save(filename)

process_wb('transactions.xlsx')





